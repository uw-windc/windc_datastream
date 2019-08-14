import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows)


def file_parser():
    path = '.{s}datasources{s}SGF{s}'.format(s=os.sep)
    file = '98states.xlsx'

    region_map = pd.read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None)
    region_map = dict(zip(region_map['from'], region_map['to']))

    sgf_map = pd.read_csv('.{s}core_maps{s}sgf.csv'.format(s=os.sep), index_col=None)
    sgf_map.drop(index=sgf_map[sgf_map['sgf_1998_relabel'].isnull() == True].index, inplace=True)

    sgf_labels = dict(zip(sgf_map['line_num_1998'], sgf_map['sgf_1998_relabel']))
    sgf_units = dict(zip(sgf_map['line_num_1998'], sgf_map['sgf_1998_units']))

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)
    a = load_workbook_range('A6:AZ56', wb[wb.sheetnames[1]])
    cols = load_workbook_range('A4:AZ4', wb[wb.sheetnames[1]])
    cols = dict(zip(cols.loc[0, :].index, cols.loc[0, :]))
    cols[0] = 'item'
    a.rename(cols, axis='columns', inplace=True)

    # strip whitespace from items
    a['item'] = [a.loc[i, 'item'].strip() for i in a.index]

    # replace cells with empty strings will null values
    a.loc[a[a['item'] == ''].index, 'item'] = None

    # drop rows that are completely empty
    a.dropna(how='all', inplace=True)

    # add in line numbering scheme
    a['line_parse'] = list(range(1, 47+1))
    a.reset_index(drop=True, inplace=True)

    # melt data
    a = pd.melt(a, id_vars=['item', 'line_parse'], var_name='region', value_name='value')

    # get rid of non-float elements
    a['value'] = pd.to_numeric(a['value'], errors='coerce')

    # zero out any null values
    a.fillna(0, inplace=True)

    # harmonize labels
    a['mapped_label'] = a['line_parse'].map(sgf_labels)
    a['units'] = a['line_parse'].map(sgf_units)
    a['region'] = a['region'].map(region_map)

    # write out a harmonized csv file with data
    df2 = a[['mapped_label', 'region', 'value', 'units']]

    return df2


def file_parser_to_csv():
    s = file_parser()
    s.to_csv('sgf_1998.csv')


if __name__ == '__main__':

    file_parser_to_csv()
