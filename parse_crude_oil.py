import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
from sql_engines import windc_engine as engine
from sqlalchemy import INTEGER, TEXT, FLOAT, text
import argparse
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows[1:], columns=data_rows[0])


def parse_file():
    path = '.{s}datasources{s}SEDS{s}CrudeOil{s}'.format(s=os.sep)
    file = 'R0000____3a.xlsx'

    region_map = pd.read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None,
                             dtype={'from': 'object', 'to': 'str'})
    region_map = dict(zip(region_map['from'], region_map['to']))

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)
    a = load_workbook_range('A3:B54', wb['Data 1'])

    a.rename({'Date': 'year', 'U.S. Crude Oil Composite Acquisition Cost by Refiners (Dollars per Barrel)': 'price'},
             axis='columns', inplace=True)

    a['units'] = 'us dollars (USD) per barrel'
    a['notes'] = 'crude oil composite acquisition cost by refiners'

    a['year'] = [a.loc[i, 'year'].year for i in a.index]

    return a


def parse_file_to_csv():
    s = parse_file()
    s.to_csv('eia_crude_oil_price.csv')


def parse_file_to_sql():
    # parse EIA Crude Oil Data (BlueNOTE) Data
    df = parse_file()

    # to sql
    types = {'year': INTEGER,
             'price': FLOAT,
             'units': TEXT,
             'notes': TEXT}

    df.to_sql('eia_crude_oil', con=engine, if_exists='replace', dtype=types)

    tbl_desc = 'EIA Annual Crude Oil Prices -- https://www.eia.gov/dnav/pet/pet_pri_rac2_dcu_nus_a.htm'
    comment = text("COMMENT ON TABLE eia_crude_oil IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--csv-out', dest='csv', action='store_true')
    parser.set_defaults(csv=False)

    args = parser.parse_args()

    if args.csv == False:
        parse_file_to_sql()
    elif args.csv == True:
        parse_file_to_csv()
