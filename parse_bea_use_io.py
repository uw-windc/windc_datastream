import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
from sql_engines import windc_engine as engine
from sqlalchemy import INTEGER, TEXT, BIGINT, text
import argparse
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows[1:], columns=data_rows[0])


def pull_data(wb, ranges, sheet_name):

    # structure is same as the original Excel file
    a = load_workbook_range(ranges['data'], wb[sheet_name])

    # row mapping
    rm = load_workbook_range(ranges['row_map'], wb[sheet_name])
    rm = dict(zip(rm['Name'], rm['IOCode']))

    # column mapping
    cm = load_workbook_range(ranges['col_map'], wb[sheet_name])
    cm = pd.DataFrame(data=cm.values.T)
    cm.drop(0, inplace=True)
    cm.reset_index(drop=True, inplace=True)
    cm = dict(zip(cm[1], cm[0]))

    # rename some stuff
    a.rename({'Name': 'industry'}, axis='columns', inplace=True)

    # need to morph data structure into a dictionary
    a = pd.melt(a, id_vars='industry', var_name='category')

    # map row and column unique ids (IOCodes)
    a['industry_code'] = a['industry'].map(rm)
    a['category_code'] = a['category'].map(cm)

    a.replace(to_replace='...', value='0', inplace=True)
    a['value'] = a['value'].map(float)
    a['units'] = 'millions of us dollars (USD)'
    a['year'] = sheet_name

    return a


def file_parser():
    path = '.{s}datasources{s}BEA{s}IO{s}'.format(s=os.sep)
    file = 'Use_SUT_Framework_1997-2017_SUM.xlsx'

    sheet_lookup = {'1997': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '1998': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '1999': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2000': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2001': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2002': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2003': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2004': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2005': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2006': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2007': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2008': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2009': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2010': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2011': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2012': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2013': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2014': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2015': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'},
                    '2016': {'data': 'B7:CP89', 'row_map': 'A7:B89', 'col_map': 'B5:CP7'}}
    # '2017':{'data':'B7:CP89', 'row_map':'A7:B89', 'col_map':'B5:CP7'}}

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)

    df = pd.DataFrame(columns=['industry', 'industry_code', 'category',
                               'category_code', 'year', 'value', 'units'])
    for k, v in sheet_lookup.items():
        df = df.append(pull_data(wb=wb, ranges=v, sheet_name=k), sort=False, ignore_index=True)

    df.rename({'industry': 'input', 'industry_code': 'input_code', 'category': 'output',
               'category_code': 'output_code'}, axis='columns', inplace=True)

    return df


def file_parser_to_csv():
    df = file_parser()
    df.to_csv('bea_use_io.csv')


def file_parser_to_sql():
    # parse BEA Use Table
    df = file_parser()

    # to sql
    types = {'input': TEXT,
             'input_code': TEXT,
             'output': TEXT,
             'output_code': TEXT,
             'year': INTEGER,
             'value': BIGINT,
             'units': TEXT}
    df.to_sql('bea_use', con=engine, if_exists='replace', dtype=types)

    tbl_desc = 'BEA IO Use Table (71 Industry) -- https://www.bea.gov/industry/input-output-accounts-data'
    comment = text("COMMENT ON TABLE bea_use IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--csv-out', dest='csv', action='store_true')
    parser.set_defaults(csv=False)

    args = parser.parse_args()

    if args.csv == False:
        file_parser_to_sql()
    elif args.csv == True:
        file_parser_to_csv()
