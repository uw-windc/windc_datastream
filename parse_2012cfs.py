from pandas import DataFrame, read_csv
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
from sql_engines import windc_engine as engine
from sqlalchemy import INTEGER, TEXT, FLOAT, text
import argparse
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return DataFrame(data_rows)


def file_parser():
    path = '.{s}datasources{s}CFS{s}'.format(s=os.sep)
    file = 'cfs_2012_pumf_csv.txt'

    region_map = read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None,
                          dtype={'from': 'object', 'to': 'str'})
    region_map = dict(zip(region_map['from'], region_map['to']))

    a = read_csv(path+file, index_col=None, low_memory=False)

    # map region names
    # must convert fips code to string to enable mapping
    a['ORIG_STATE'] = a['ORIG_STATE'].map(str)
    a['DEST_STATE'] = a['DEST_STATE'].map(str)

    a['ORIG_STATE'] = a['ORIG_STATE'].map(region_map)
    a['DEST_STATE'] = a['DEST_STATE'].map(region_map)

    # open supplemental data file
    sup_file = 'cfs_2012_pum_file_users_guide_App_A (Jun 2015).xlsx'
    wb = load_workbook(filename=path+sup_file, read_only=True, data_only=True)

    # adding CFS Area descriptions from supplemental data files
    apdx1 = load_workbook_range('C2:E136', wb['App A1'])
    apdx1.drop(index=[0, 1], inplace=True)
    apdx1.reset_index(drop=True, inplace=True)
    apdx1.rename({0: 'from', 2: 'to'}, axis='columns', inplace=True)
    apdx1.drop(columns=[1], inplace=True)

    # a['ORIG_CFS_AREA_desc'] = a['ORIG_CFS_AREA'].map(dict(zip(apdx1[0],apdx1[2])))
    # a['DEST_CFS_AREA_desc'] = a['DEST_CFS_AREA'].map(dict(zip(apdx1[0],apdx1[2])))

    # adding NAICS descriptions from supplemental data files
    apdx2 = load_workbook_range('A2:B47', wb['App A2'])
    apdx2.drop(index=[0], inplace=True)
    apdx2.reset_index(drop=True, inplace=True)
    apdx2.rename({0: 'from', 1: 'to'}, axis='columns', inplace=True)

    # a['NAICS_desc'] = a['NAICS'].map(dict(zip(apdx2[0],apdx2[1])))

    # adding SCTG descriptions from supplemental data files
    apdx3 = load_workbook_range('A2:C46', wb['App A3'])
    apdx3.drop(index=[0], inplace=True)
    apdx3[0] = apdx3[0].map(str)

    gmap = dict(zip(apdx3[0], apdx3[1]))
    a['SCTG_desc'] = a['SCTG'].map(gmap)

    group_map = {}
    g = list(apdx3.loc[apdx3[2].isnull() == False, 2])
    for i in g:
        idx_start = apdx3[apdx3[0] == i.split('-')[0]].index[0]
        idx_end = apdx3[apdx3[0] == i.split('-')[1]].index[0]

        gg = list(apdx3.loc[idx_start:idx_end, 1])

        label = ''
        for j in gg:
            label = label + str(j) + ', '

        group_map[i] = label

    gmap.update(group_map)

    apdx3 = DataFrame(columns=['from', 'to'])
    apdx3['from'] = gmap.keys()
    apdx3['to'] = apdx3['from'].map(gmap)

    # a['SCTG_desc'] = a['SCTG'].map(gmap)

    # adding Mode descriptions from supplemental data files
    apdx4 = load_workbook_range('A2:B23', wb['App A4'])
    apdx4.drop(index=[0], inplace=True)
    apdx4[0] = apdx4[0].map(int)
    apdx4[1] = [apdx4.loc[i, 1].strip() for i in apdx4.index]

    apdx4.rename({0: 'from', 1: 'to'}, axis='columns', inplace=True)

    # a['MODE_desc'] = a['MODE'].map(dict(zip(apdx4[0],apdx4[1])))

    # adding HAZMAT description from supplemental data files
    hazmat = {'P': 'Class 3 HAZMAT (flammable liquids)', 'H': 'Other HAZMAT', 'N': 'Not HAZMAT'}
    a['HAZMAT'] = a['HAZMAT'].map(hazmat)

    # Export country
    cntry = {'C': 'Canada', 'M': 'Mexico', 'O': 'Other'}
    a['EXPORT_CNTRY'] = a['EXPORT_CNTRY'].map(cntry)

    # adding in units
    a['SHIPMT_VALUE_units'] = 'us dollars (USD)'
    a['SHIPMT_WGHT_units'] = 'weight of shipment in pounds'
    a['SHIPMT_DIST_GC_units'] = 'great circle distance in miles'
    a['SHIPMT_DIST_ROUTED_units'] = 'routed distance in miles'

    # cols = ['SHIPMT_ID','ORIG_STATE','ORIG_MA','ORIG_CFS_AREA','ORIG_CFS_AREA_desc','DEST_STATE','DEST_MA','DEST_CFS_AREA','DEST_CFS_AREA_desc','NAICS','NAICS_desc','QUARTER','SCTG','SCTG_desc','MODE','MODE_desc','SHIPMT_VALUE','SHIPMT_VALUE_units','SHIPMT_WGHT','SHIPMT_WGHT_units','SHIPMT_DIST_GC','SHIPMT_DIST_GC_units','SHIPMT_DIST_ROUTED','SHIPMT_DIST_ROUTED_units','TEMP_CNTL_YN','EXPORT_YN','EXPORT_CNTRY','HAZMAT','WGT_FACTOR']

    return a, apdx1, apdx2, apdx3, apdx4


def file_parser_to_csv():
    a, apdx1, apdx2, apdx3, apdx4 = file_parser()
    a.to_csv('cfs_2012_pumf.csv')
    apdx1.to_csv('cfs_apdx1.csv')
    apdx2.to_csv('cfs_apdx2.csv')
    apdx3.to_csv('cfs_apdx3.csv')
    apdx4.to_csv('cfs_apdx4.csv')


def file_parser_to_sql():
    # parse Commodity Flow Survey
    df, apdx1, apdx2, apdx3, apdx4 = file_parser()

    types = {'SHIPMT_ID': INTEGER,
             'ORIG_STATE': TEXT,
             'ORIG_MA': TEXT,
             'ORIG_CFS_AREA': TEXT,
             'DEST_STATE': TEXT,
             'DEST_MA': TEXT,
             'DEST_CFS_AREA': TEXT,
             'NAICS': TEXT,
             'QUARTER': INTEGER,
             'SCTG': TEXT,
             'SCTG_desc': TEXT,
             'MODE': INTEGER,
             'SHIPMT_VALUE': FLOAT,
             'SHIPMT_VALUE_units': TEXT,
             'SHIPMT_WGHT': FLOAT,
             'SHIPMT_WGHT_units': TEXT,
             'SHIPMT_DIST_GC': INTEGER,
             'SHIPMT_DIST_GC_units': TEXT,
             'SHIPMT_DIST_ROUTED': INTEGER,
             'SHIPMT_DIST_ROUTED_units': TEXT,
             'TEMP_CNTL_YN': TEXT,
             'EXPORT_YN': TEXT,
             'EXPORT_CNTRY': TEXT,
             'HAZMAT': TEXT,
             'WGT_FACTOR': FLOAT}
    df.to_sql('cfs_2012', con=engine, if_exists='replace', dtype=types)

    tbl_desc = '2012 Commodity Flow Survey -- https://www.census.gov/programs-surveys/cfs.html'
    comment = text("COMMENT ON TABLE cfs_2012 IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)

    types = {'from': TEXT, 'to': TEXT}
    apdx1.to_sql('cfs_2012_orig_dest_cfs_area_map', con=engine, if_exists='replace', dtype=types)

    tbl_desc = '2012 Commodity Flow Survey -- Descriptions from Appendix 1'
    comment = text("COMMENT ON TABLE cfs_2012_orig_dest_cfs_area_map IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)

    types = {'from': INTEGER, 'to': TEXT}
    apdx2['from'] = apdx2['from'].map(int)
    apdx2.to_sql('cfs_2012_naics_map', con=engine, if_exists='replace', dtype=types)

    tbl_desc = '2012 Commodity Flow Survey -- Descriptions from Appendix 2'
    comment = text("COMMENT ON TABLE cfs_2012_naics_map IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)

    types = {'from': TEXT, 'to': TEXT}
    apdx3.to_sql('cfs_2012_sctg_map', con=engine, if_exists='replace')

    tbl_desc = '2012 Commodity Flow Survey -- Descriptions from Appendix 3'
    comment = text("COMMENT ON TABLE cfs_2012_sctg_map IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)

    types = {'from': INTEGER, 'to': TEXT}
    apdx4['from'] = apdx4['from'].map(int)
    apdx4.to_sql('cfs_2012_mode_map', con=engine, if_exists='replace')

    tbl_desc = '2012 Commodity Flow Survey -- Descriptions from Appendix 4'
    comment = text("COMMENT ON TABLE cfs_2012_mode_map IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--csv-out', dest='csv', action='store_true')
    parser.set_defaults(csv=False)

    args = parser.parse_args()

    if args.csv == False:
        file_parser_to_sql()
    elif args.csv == True:
        file_parser_to_csv()
