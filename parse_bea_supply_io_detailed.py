import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
from sql_engines import windc_engine as engine
from sqlalchemy import INTEGER, TEXT, FLOAT, text
import argparse
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows)


def pull_data(sheet_name, data, cols, rows):
    path = '.{s}datasources{s}BEA_2007_2012{s}'.format(s=os.sep)
    file = 'Supply_2007_2012_DET.xlsx'

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)
    data = load_workbook_range(data, wb[sheet_name])

    # coerce data to numeric values
    for i in data.keys():
        data[i] = pd.to_numeric(data[i], errors='coerce')

    # column descriptive data
    b = load_workbook_range(cols, wb[sheet_name])
    b = {str(b.loc[1, i]): b.loc[0, i].strip() for i in b.keys()}

    # row descriptive data
    c = load_workbook_range(rows, wb[sheet_name])
    c = {str(c.loc[i, 0]): c.loc[i, 1].strip() for i in c.index}

    # rename columns
    cols = {k: v for k, v in enumerate(b.keys())}
    data.rename(cols, axis='columns', inplace=True)

    # add row labels
    data['rows'] = c.keys()

    # melt data
    a = pd.melt(data, id_vars=['rows'], var_name='to_industry')
    a.rename({'rows': 'from_industry'}, axis='columns', inplace=True)
    a['units'] = 'millions of us dollars (USD)'

    # merge b and c dictionary
    b.update(c)

    a['from_industry_desc'] = a['from_industry'].map(b)
    a['to_industry_desc'] = a['to_industry'].map(b)
    a['year'] = int(sheet_name)

    # replace null values with zeros
    a['value'].fillna(0, inplace=True)

    # organize columns a bit
    a = a[['from_industry', 'from_industry_desc', 'to_industry',
           'to_industry_desc', 'year', 'value', 'units']]

    return a


def file_parser():
    sheet_lookup = {'2007': {'data': 'C7:PC412', 'cols': 'C5:PC6', 'rows': 'A7:B412'},
                    '2012': {'data': 'C7:PC412', 'cols': 'C5:PC6', 'rows': 'A7:B412'}}

    df = pd.DataFrame()
    for i in sheet_lookup.keys():
        a = pull_data(sheet_name=i, data=sheet_lookup[i]['data'],
                      cols=sheet_lookup[i]['cols'], rows=sheet_lookup[i]['rows'])
        df = df.append(a, sort=False, ignore_index=True)

    return df


def file_parser_to_csv():
    df = file_parser()
    df.to_csv('bea_supply_io_detailed.csv')


def file_parser_to_sql():
    df = file_parser()

    # to sql
    types = {'from_industry': TEXT,
             'to_industry': TEXT,
             'from_industry_desc': TEXT,
             'to_industry_desc': TEXT,
             'year': INTEGER,
             'value': FLOAT,
             'units': TEXT}
    df.to_sql('bea_supply_detailed', con=engine, if_exists='replace', dtype=types)

    tbl_desc = 'BEA IO Supply Table (Detailed) -- https://www.bea.gov/industry/input-output-accounts-data'
    comment = text("COMMENT ON TABLE bea_supply_detailed IS '{}';".format(
        tbl_desc)).execution_options(autocommit=True)
    engine.execute(comment)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--csv-out', dest='csv', action='store_true')
    parser.set_defaults(csv=False)

    args = parser.parse_args()

    if args.csv == False:
        file_parser_to_sql()
    elif args.csv == True:
        file_parser_to_csv()
