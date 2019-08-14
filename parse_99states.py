import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows)


def file_parser():
    path = '.{s}datasources{s}SGF{s}'.format(s=os.sep)
    file = '99statess.xlsx'

    region_map = pd.read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None)
    region_map = dict(zip(region_map['from'], region_map['to']))

    sgf_map = pd.read_csv('.{s}core_maps{s}sgf.csv'.format(s=os.sep), index_col=None)
    sgf_map.drop(index=sgf_map[sgf_map['sgf_1999_relabel'].isnull() == True].index, inplace=True)
    sgf_labels = dict(zip(sgf_map['line_num_1999'], sgf_map['sgf_1999_relabel']))
    sgf_units = dict(zip(sgf_map['line_num_1999'], sgf_map['sgf_1999_units']))

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)

    # this gets the first part of the data... need to pull in sheet 2
    a = load_workbook_range('A8:BX64', wb[wb.sheetnames[0]])
    cols = load_workbook_range('A3:BX3', wb[wb.sheetnames[0]])
    cols = dict(zip(cols.loc[0, :].index, cols.loc[0, :]))
    cols[0] = 'item'
    cols = {k: cols[k] if cols[k] != None else 'drop_me' for k in cols.keys()}
    a.rename(cols, axis='columns', inplace=True)
    a.drop(columns='drop_me', inplace=True)

    # pulling in data from sheet 2
    b = load_workbook_range('A8:CA64', wb[wb.sheetnames[1]])
    cols = load_workbook_range('A3:CA3', wb[wb.sheetnames[1]])
    cols = dict(zip(cols.loc[0, :].index, cols.loc[0, :]))
    cols[0] = 'item'
    cols = {k: cols[k] if cols[k] != None else 'drop_me' for k in cols.keys()}
    b.rename(cols, axis='columns', inplace=True)
    b.drop(columns=['item', 'drop_me'], inplace=True)

    # merge datasets
    c = a.copy()
    for i in b.keys():
        c[i] = b[i]

    # drop rows that are completely empty
    c.dropna(how='all', inplace=True)
    c.reset_index(drop=True, inplace=True)

    # strip whitespace from items
    c['item'] = [c.loc[i, 'item'].strip() for i in c.index]

    # add in line numbering scheme
    c['line_parse'] = list(range(1, 48+1))

    # melt data
    c = pd.melt(c, id_vars=['item', 'line_parse'], var_name='region', value_name='value')

    # get rid of non-float elements
    c['value'] = pd.to_numeric(c['value'], errors='coerce')

    # zero out any null values
    c.fillna(0, inplace=True)

    # harmonize labels
    c['mapped_label'] = c['line_parse'].map(sgf_labels)
    c['units'] = c['line_parse'].map(sgf_units)
    c['region'] = c['region'].map(region_map)

    # write out a harmonized csv file with data
    df2 = c[['mapped_label', 'region', 'value', 'units']]

    return df2


def file_parser_to_csv():
    s = file_parser()
    s.to_csv('sgf_1999.csv')


if __name__ == '__main__':

    file_parser_to_csv()
