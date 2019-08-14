import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows)


def parse_file(filename, year, data_rng, col_rng):
    path = '.{s}datasources{s}SGF{s}'.format(s=os.sep)
    file = filename

    region_map = pd.read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None)
    region_map = dict(zip(region_map['from'], region_map['to']))

    sgf_map = pd.read_csv('.{s}core_maps{s}sgf.csv'.format(s=os.sep), index_col=None)
    sgf_map.drop(index=sgf_map[sgf_map['sgf_'+str(year) +
                                       '_relabel'].isnull() == True].index, inplace=True)

    sgf_labels = dict(zip(sgf_map['line_num_'+str(year)], sgf_map['sgf_'+str(year)+'_relabel']))
    sgf_units = dict(zip(sgf_map['line_num_'+str(year)], sgf_map['sgf_'+str(year)+'_units']))

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)

    a = load_workbook_range(data_rng, wb[wb.sheetnames[0]])
    cols = load_workbook_range(col_rng, wb[wb.sheetnames[0]])
    cols = dict(zip(cols.loc[0, :].index, cols.loc[0, :]))
    cols[0] = 'item'
    cols = {k: cols[k] if cols[k] != None else 'drop_me' for k in cols.keys()}
    a.rename(cols, axis='columns', inplace=True)

    if 'drop_me' in a.keys():
        a.drop(columns='drop_me', inplace=True)

    # drop rows that are completely empty
    a.dropna(how='all', inplace=True)
    a.reset_index(drop=True, inplace=True)

    # strip whitespace from items
    a['item'] = [a.loc[i, 'item'].strip() for i in a.index]

    # add in line numbering scheme
    a['line_parse'] = list(range(1, len(a)+1))

    # melt data
    a = pd.melt(a, id_vars=['item', 'line_parse'], var_name='region', value_name='value')

    # get rid of non-float elements
    a['value'] = pd.to_numeric(a['value'], errors='coerce')

    # zero out any null values
    a.loc[a[a['value'].isnull() == True].index, 'value'] = 0

    # harmonize labels
    a['mapped_label'] = a['line_parse'].map(sgf_labels)
    a['units'] = a['line_parse'].map(sgf_units)
    a['region'] = a['region'].map(region_map)

    # remove all rows that do not have a mapped label
    # this step de-dups as well as applies a consistent labeling convention
    a.drop(a[a['mapped_label'].isnull() == True].index, inplace=True)
    a.reset_index(drop=True, inplace=True)

    # write out a harmonized csv file with data
    df2 = a[['mapped_label', 'region', 'value', 'units']]

    return df2


def parse_file_to_csv(filename, year, data_rng, col_rng):
    a = parse_file(filename=filename, year=year, data_rng=data_rng, col_rng=col_rng)
    a.to_csv('sgf_'+str(year)+'.csv')


def file_parser_to_csv():
    t = {'00statess.xlsx': {'year': '2000', 'data_rng': 'A8:EX64', 'col_rng': 'A3:EX3'},
         '01statess.xlsx': {'year': '2001', 'data_rng': 'A8:EX66', 'col_rng': 'A3:EX3'},
         '02statess.xlsx': {'year': '2002', 'data_rng': 'A8:CY66', 'col_rng': 'A3:CY3'},
         '03statess.xlsx': {'year': '2003', 'data_rng': 'A8:CY65', 'col_rng': 'A3:CY3'},
         '04statess.xlsx': {'year': '2004', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '05statess.xlsx': {'year': '2005', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '06statess.xlsx': {'year': '2006', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '07statess.xlsx': {'year': '2007', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '08statess.xlsx': {'year': '2008', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '09statess.xlsx': {'year': '2009', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '10statess.xlsx': {'year': '2010', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'},
         '11statess.xlsx': {'year': '2011', 'data_rng': 'A8:AZ64', 'col_rng': 'A3:AZ3'}}

    for i in t.keys():
        parse_file_to_csv(filename=i, year=t[i]['year'],
                          data_rng=t[i]['data_rng'], col_rng=t[i]['col_rng'])


if __name__ == '__main__':

    file_parser_to_csv()
