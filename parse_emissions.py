import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
import argparse
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(columns=data_rows[0], data=data_rows[1:])


def parse_file(filename, data_rng, nickname):
    path = '.{s}datasources{s}SEDS{s}Emissions{s}'.format(s=os.sep)
    file = filename

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)
    a = load_workbook_range(data_rng, wb[wb.sheetnames[0]])

    a.rename({'State': 'region'}, axis='columns', inplace=True)

    # strip whitespace from items
    a['region'] = [a.loc[i, 'region'].strip() for i in a.index]

    # harmonize labels
    region_map = pd.read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None)
    region_map = dict(zip(region_map['from'], region_map['to']))

    a['region'] = a['region'].map(region_map)

    # melt data
    a = pd.melt(a, id_vars=['region'], var_name='year', value_name='emissions')

    a['units'] = 'million metric tons of carbon dioxide'

    return a


def parse_file_to_csv(filename, data_rng, nickname):
    s = parse_file(filename=filename, data_rng=data_rng, nickname=nickname)
    s.to_csv('seds_'+nickname+'.csv')


def file_parser_to_csv():
    t = {'coal_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'coal'},
         'natural_gas_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'natgas'},
         'petroleum_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'petrol'},
         'industrial_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'ind'},
         'commercial_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'com'},
         'residential_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'res'},
         'electric_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'elec'},
         'transportation_CO2_by_state_2013.xlsx': {'data_rng': 'A3:AI55', 'nickname': 'trans'}}

    for i in t.keys():
        parse_file_to_csv(filename=i, data_rng=t[i]['data_rng'], nickname=t[i]['nickname'])


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--csv-out', dest='csv', action='store_true')
    parser.set_defaults(csv=False)

    args = parser.parse_args()

    if args.csv == False:
        file_parser_to_sql()
    elif args.csv == True:
        file_parser_to_csv()
