import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_interval
import re
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows)


def file_parser():
    path = '.{s}datasources{s}SGF{s}'.format(s=os.sep)
    file = '97states.xlsx'

    region_map = pd.read_csv('.{s}core_maps{s}regions.csv'.format(s=os.sep), index_col=None)
    region_map = dict(zip(region_map['from'], region_map['to']))

    sgf_map = pd.read_csv('.{s}core_maps{s}sgf.csv'.format(s=os.sep), index_col=None)
    sgf_map.drop(index=sgf_map[sgf_map['sgf_1997_relabel'].isnull() == True].index, inplace=True)
    sgf_labels = dict(zip(sgf_map['line_num_1997'], sgf_map['sgf_1997_relabel']))

    sgf_units = dict(zip(sgf_map['line_num_1997'], sgf_map['sgf_1997_units']))

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)
    a = load_workbook_range('A7:D2759', wb[wb.sheetnames[0]])

    a.dropna(how='all', inplace=True)
    a.rename({0: 'item', 1: 'value', 2: 'Percent distribution',
              3: 'Per capita'}, axis='columns', inplace=True)
    a['item'] = [a.loc[i, 'item'].strip() for i in a.index]
    a.drop(columns=['Percent distribution', 'Per capita'], inplace=True)
    a.reset_index(drop=True, inplace=True)

    # find row with a region
    a['isRegion'] = [a.loc[i, 'item'] in region_map.keys() for i in range(len(a))]
    region_parse = dict(zip(a[a['isRegion'] == True].item, a[a['isRegion'] == True].index))
    found_regions = list(a.loc[a['isRegion'] == True, 'item'])

    # add in line numbering scheme to aid in parsing
    a['line_parse'] = 0
    for i in range(len(region_parse)):
        if i < len(region_parse) - 1:
            idx = a.loc[region_parse[found_regions[i]] +
                        1:region_parse[found_regions[i+1]]-1, :].index
            a.loc[idx, 'line_parse'] = list(range(1, 47+1))
        else:
            idx = a.loc[region_parse[found_regions[i]]+1:len(a)-1, :].index
            a.loc[idx, 'line_parse'] = list(range(1, 47+1))

    a['mapped_label'] = a['line_parse'].map(sgf_labels)
    a['units'] = a['line_parse'].map(sgf_units)

    df = pd.DataFrame(columns=['region', 'value', 'units'])
    for i in range(len(region_parse)):
        if i < len(region_parse) - 1:
            # just a staging dataframe to append
            d = a.loc[region_parse[found_regions[i]]+1:region_parse[found_regions[i+1]]-1, :].copy()
            d['region'] = found_regions[i]

            # append to dataframe
            df = df.append(d, sort=False, ignore_index=True)
        else:
            d = a.loc[region_parse[found_regions[i]]+1:len(a)-1, :].copy()
            d['region'] = found_regions[i]

            # append to dataframe
            df = df.append(d, sort=False, ignore_index=True)

    # harmonize region names
    df['region'] = df['region'].map(region_map)

    # fill in null values with numeric zero
    df.fillna(0, inplace=True)

    # write out a harmonized csv file with data
    df2 = df[['mapped_label', 'region', 'value', 'units']]

    return df2


def file_parser_to_csv():
    s = file_parser()
    s.to_csv('sgf_1997.csv')


if __name__ == '__main__':

    file_parser_to_csv()
